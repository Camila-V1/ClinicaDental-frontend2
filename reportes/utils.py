"""
Utilidades para generación de reportes en diferentes formatos
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class PDFReportGenerator:
    """Generador de reportes PDF con formato profesional"""
    
    def __init__(self, title, tenant_name="Clínica Dental"):
        self.buffer = BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=1*inch,
            bottomMargin=0.75*inch
        )
        self.title = title
        self.tenant_name = tenant_name
        self.styles = getSampleStyleSheet()
        self.story = []
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        ))
    
    def add_header(self):
        """Añade encabezado al reporte"""
        # Título de la clínica
        clinic_title = Paragraph(self.tenant_name, self.styles['CustomTitle'])
        self.story.append(clinic_title)
        
        # Título del reporte
        report_title = Paragraph(self.title, self.styles['CustomSubtitle'])
        self.story.append(report_title)
        
        # Fecha de generación
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        fecha_text = Paragraph(f"<i>Generado el: {fecha}</i>", self.styles['CustomNormal'])
        self.story.append(fecha_text)
        self.story.append(Spacer(1, 0.3*inch))
    
    def add_table(self, data, col_widths=None, title=None):
        """
        Añade una tabla al reporte
        
        Args:
            data: Lista de listas con los datos (primera fila = encabezados)
            col_widths: Lista con anchos de columnas (opcional)
            title: Título de la tabla (opcional)
        """
        if title:
            self.story.append(Paragraph(title, self.styles['CustomSubtitle']))
            self.story.append(Spacer(1, 0.1*inch))
        
        if not data or len(data) == 0:
            self.story.append(Paragraph("No hay datos disponibles", self.styles['CustomNormal']))
            return
        
        # Crear tabla
        table = Table(data, colWidths=col_widths)
        
        # Estilo de la tabla
        style = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e3a8a')),
            
            # Alternar colores en filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        table.setStyle(style)
        self.story.append(table)
        self.story.append(Spacer(1, 0.3*inch))
    
    def add_key_metrics(self, metrics):
        """
        Añade métricas clave en formato destacado
        
        Args:
            metrics: Diccionario con nombre_metrica: valor
        """
        data = [['Métrica', 'Valor']]
        for key, value in metrics.items():
            data.append([key, str(value)])
        
        self.add_table(data, col_widths=[4*inch, 2*inch], title="Métricas Principales")
    
    def add_paragraph(self, text, style='CustomNormal'):
        """Añade un párrafo de texto"""
        self.story.append(Paragraph(text, self.styles[style]))
        self.story.append(Spacer(1, 0.1*inch))
    
    def generate(self):
        """Genera el PDF y retorna HttpResponse"""
        self.doc.build(self.story)
        self.buffer.seek(0)
        
        response = HttpResponse(self.buffer.getvalue(), content_type='application/pdf')
        filename = f"{self.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ExcelReportGenerator:
    """Generador de reportes Excel con formato profesional"""
    
    def __init__(self, title, tenant_name="Clínica Dental"):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.title = title
        self.tenant_name = tenant_name
        self.current_row = 1
        self._setup_styles()
    
    def _setup_styles(self):
        """Define estilos reutilizables"""
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        
        self.title_font = Font(name='Arial', size=16, bold=True, color='1e3a8a')
        self.subtitle_font = Font(name='Arial', size=12, bold=True, color='3b82f6')
        self.normal_font = Font(name='Arial', size=10)
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_header(self):
        """Añade encabezado al reporte"""
        # Título de la clínica
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.worksheet[f'A{self.current_row}']
        cell.value = self.tenant_name
        cell.font = self.title_font
        cell.alignment = self.center_alignment
        self.current_row += 1
        
        # Título del reporte
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.worksheet[f'A{self.current_row}']
        cell.value = self.title
        cell.font = self.subtitle_font
        cell.alignment = self.center_alignment
        self.current_row += 1
        
        # Fecha
        self.worksheet.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.worksheet[f'A{self.current_row}']
        cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        cell.font = Font(name='Arial', size=9, italic=True)
        cell.alignment = self.center_alignment
        self.current_row += 2
    
    def add_table(self, data, title=None):
        """
        Añade una tabla al reporte
        
        Args:
            data: Lista de listas (primera fila = encabezados)
            title: Título de la tabla (opcional)
        """
        if title:
            self.worksheet.merge_cells(
                f'A{self.current_row}:{get_column_letter(len(data[0]))}{self.current_row}'
            )
            cell = self.worksheet[f'A{self.current_row}']
            cell.value = title
            cell.font = self.subtitle_font
            cell.alignment = self.left_alignment
            self.current_row += 1
        
        if not data or len(data) == 0:
            cell = self.worksheet[f'A{self.current_row}']
            cell.value = "No hay datos disponibles"
            cell.font = self.normal_font
            self.current_row += 2
            return
        
        # Encabezados
        for col_idx, header in enumerate(data[0], start=1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        self.current_row += 1
        
        # Datos
        for row_data in data[1:]:
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.left_alignment
                cell.border = self.border
                
                # Alternar color de fondo
                if self.current_row % 2 == 0:
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            self.current_row += 1
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(data[0]) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(self.worksheet[f'{column_letter}{row}'].value or ''))
                for row in range(self.current_row - len(data), self.current_row)
            )
            self.worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        self.current_row += 1
    
    def add_key_metrics(self, metrics):
        """
        Añade métricas clave en formato destacado
        
        Args:
            metrics: Diccionario con nombre_metrica: valor
        """
        data = [['Métrica', 'Valor']]
        for key, value in metrics.items():
            data.append([key, value])
        
        self.add_table(data, title="Métricas Principales")
    
    def generate(self):
        """Genera el Excel y retorna HttpResponse"""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{self.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


def format_currency(value):
    """Formatea un valor como moneda"""
    if value is None:
        return "$0.00"
    return f"${value:,.2f}"


def format_percentage(value, decimals=2):
    """Formatea un valor como porcentaje"""
    if value is None:
        return "0%"
    return f"{value:.{decimals}f}%"


def format_date(date_obj):
    """Formatea una fecha"""
    if date_obj is None:
        return "N/A"
    if isinstance(date_obj, str):
        return date_obj
    return date_obj.strftime("%d/%m/%Y")
